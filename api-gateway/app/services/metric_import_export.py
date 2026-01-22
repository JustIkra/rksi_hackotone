"""
Service for importing and exporting metric definitions.

Supports Excel (.xlsx) and JSON formats for bulk metric management.
"""

import json
from decimal import Decimal
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.ext.asyncio import AsyncSession

from app.repositories.metric import MetricDefRepository
from app.repositories.metric_category import MetricCategoryRepository
from app.schemas.metric_import import (
    ExportMetricItem,
    ImportError,
    ImportPreviewItem,
    ImportPreviewResponse,
    ImportResultResponse,
)


class MetricImportExportService:
    """Service for importing and exporting metric definitions."""

    # Excel column mapping
    XLSX_COLUMNS = ["code", "name", "name_ru", "description", "unit", "min_value", "max_value", "active", "category_code"]
    XLSX_HEADERS = ["Code", "Name", "Name (RU)", "Description", "Unit", "Min Value", "Max Value", "Active", "Category Code"]

    def __init__(self, db: AsyncSession):
        self.db = db
        self.metric_repo = MetricDefRepository(db)
        self.category_repo = MetricCategoryRepository(db)

    async def export_xlsx(self) -> bytes:
        """
        Export all metrics to Excel format.

        Returns:
            bytes: Excel file content
        """
        metrics = await self.metric_repo.list_all(active_only=False)

        # Build category_id -> code mapping
        categories = await self.category_repo.list_all()
        category_code_map = {cat.id: cat.code for cat in categories}

        wb = Workbook()
        ws = wb.active
        ws.title = "Metrics"

        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Write headers
        for col_idx, header in enumerate(self.XLSX_HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Write data
        for row_idx, metric in enumerate(metrics, start=2):
            ws.cell(row=row_idx, column=1, value=metric.code)
            ws.cell(row=row_idx, column=2, value=metric.name)
            ws.cell(row=row_idx, column=3, value=metric.name_ru)
            ws.cell(row=row_idx, column=4, value=metric.description)
            ws.cell(row=row_idx, column=5, value=metric.unit)
            ws.cell(row=row_idx, column=6, value=float(metric.min_value) if metric.min_value is not None else None)
            ws.cell(row=row_idx, column=7, value=float(metric.max_value) if metric.max_value is not None else None)
            ws.cell(row=row_idx, column=8, value="true" if metric.active else "false")
            # Add category_code
            category_code = category_code_map.get(metric.category_id) if metric.category_id else None
            ws.cell(row=row_idx, column=9, value=category_code)

        # Adjust column widths
        column_widths = [15, 30, 30, 50, 10, 12, 12, 10, 20]
        for col_idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

        # Freeze header row
        ws.freeze_panes = "A2"

        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    async def export_json(self) -> dict[str, Any]:
        """
        Export all metrics to JSON format.

        Returns:
            dict: JSON-serializable dictionary with metrics
        """
        metrics = await self.metric_repo.list_all(active_only=False)

        # Build category_id -> code mapping
        categories = await self.category_repo.list_all()
        category_code_map = {cat.id: cat.code for cat in categories}

        items = []
        for metric in metrics:
            category_code = category_code_map.get(metric.category_id) if metric.category_id else None
            items.append(
                ExportMetricItem(
                    code=metric.code,
                    name=metric.name,
                    name_ru=metric.name_ru,
                    description=metric.description,
                    unit=metric.unit,
                    min_value=float(metric.min_value) if metric.min_value is not None else None,
                    max_value=float(metric.max_value) if metric.max_value is not None else None,
                    active=metric.active,
                    category_code=category_code,
                ).model_dump()
            )

        return {"metrics": items, "total": len(items)}

    async def import_preview(self, file_content: bytes, filename: str) -> ImportPreviewResponse:
        """
        Preview import changes without applying them.

        Args:
            file_content: Raw file content
            filename: Original filename (used to detect format)

        Returns:
            ImportPreviewResponse with to_create, to_update, and errors
        """
        # Parse file
        if filename.lower().endswith(".xlsx"):
            parsed_metrics, parse_errors = self._parse_xlsx(file_content)
        elif filename.lower().endswith(".json"):
            parsed_metrics, parse_errors = self._parse_json(file_content)
        else:
            return ImportPreviewResponse(
                errors=[ImportError(row=0, error=f"Unsupported file format: {filename}")]
            )

        # Get existing metrics by code
        existing_metrics = await self.metric_repo.list_all(active_only=False)
        existing_by_code = {m.code: m for m in existing_metrics}

        # Build category_id -> code mapping for change detection
        categories = await self.category_repo.list_all()
        category_code_map = {cat.id: cat.code for cat in categories}

        to_create: list[ImportPreviewItem] = []
        to_update: list[ImportPreviewItem] = []
        errors: list[ImportError] = list(parse_errors)

        for idx, metric_data in enumerate(parsed_metrics, start=2):
            code = metric_data.get("code")
            if not code:
                errors.append(ImportError(row=idx, error="Missing required field: code"))
                continue

            existing = existing_by_code.get(code)

            if existing:
                # Get current category code for change detection
                existing_category_code = (
                    category_code_map.get(existing.category_id)
                    if existing.category_id
                    else None
                )
                # Check for changes
                changes = self._detect_changes(existing, metric_data, existing_category_code)
                if changes:
                    to_update.append(
                        ImportPreviewItem(
                            code=code,
                            name=metric_data.get("name"),
                            changes=changes,
                        )
                    )
            else:
                to_create.append(
                    ImportPreviewItem(
                        code=code,
                        name=metric_data.get("name"),
                        changes=None,
                    )
                )

        return ImportPreviewResponse(to_create=to_create, to_update=to_update, errors=errors)

    async def import_metrics(self, file_content: bytes, filename: str) -> ImportResultResponse:
        """
        Import metrics from file (upsert mode).

        Args:
            file_content: Raw file content
            filename: Original filename (used to detect format)

        Returns:
            ImportResultResponse with created, updated counts and errors
        """
        # Parse file
        if filename.lower().endswith(".xlsx"):
            parsed_metrics, parse_errors = self._parse_xlsx(file_content)
        elif filename.lower().endswith(".json"):
            parsed_metrics, parse_errors = self._parse_json(file_content)
        else:
            return ImportResultResponse(
                created=0,
                updated=0,
                errors=[ImportError(row=0, error=f"Unsupported file format: {filename}")],
            )

        # Get existing metrics by code
        existing_metrics = await self.metric_repo.list_all(active_only=False)
        existing_by_code = {m.code: m for m in existing_metrics}

        created = 0
        updated = 0
        errors: list[ImportError] = list(parse_errors)

        # Cache for resolved categories to avoid duplicate lookups/creates
        category_cache: dict[str, Any] = {}

        for idx, metric_data in enumerate(parsed_metrics, start=2):
            code = metric_data.get("code")
            if not code:
                errors.append(ImportError(row=idx, error="Missing required field: code"))
                continue

            name = metric_data.get("name")
            if not name:
                errors.append(ImportError(row=idx, error=f"Missing required field: name for code '{code}'"))
                continue

            try:
                # Resolve category_code to category_id
                category_code = metric_data.get("category_code")
                category_id = await self._resolve_category_id(category_code, category_cache)

                existing = existing_by_code.get(code)

                if existing:
                    # Update existing
                    await self.metric_repo.update(
                        metric_def_id=existing.id,
                        name=name,
                        name_ru=metric_data.get("name_ru"),
                        description=metric_data.get("description"),
                        unit=metric_data.get("unit"),
                        min_value=self._to_decimal(metric_data.get("min_value")),
                        max_value=self._to_decimal(metric_data.get("max_value")),
                        active=metric_data.get("active", True),
                        category_id=category_id,
                    )
                    updated += 1
                else:
                    # Create new
                    await self.metric_repo.create(
                        code=code,
                        name=name,
                        name_ru=metric_data.get("name_ru"),
                        description=metric_data.get("description"),
                        unit=metric_data.get("unit"),
                        min_value=self._to_decimal(metric_data.get("min_value")),
                        max_value=self._to_decimal(metric_data.get("max_value")),
                        active=metric_data.get("active", True),
                        category_id=category_id,
                    )
                    created += 1

            except Exception as e:
                errors.append(ImportError(row=idx, error=f"Failed to save metric '{code}': {str(e)}"))

        return ImportResultResponse(created=created, updated=updated, errors=errors)

    def _parse_xlsx(self, content: bytes) -> tuple[list[dict[str, Any]], list[ImportError]]:
        """
        Parse Excel file content.

        Args:
            content: Excel file bytes

        Returns:
            Tuple of (parsed metrics list, parsing errors)
        """
        metrics: list[dict[str, Any]] = []
        errors: list[ImportError] = []

        try:
            wb = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
            ws = wb.active

            if ws is None:
                errors.append(ImportError(row=0, error="Excel file has no active worksheet"))
                return metrics, errors

            # Get headers from first row
            header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            if not header_row:
                errors.append(ImportError(row=1, error="Empty header row"))
                return metrics, errors

            # Map headers to column indices (case-insensitive)
            header_map = {}
            for col_idx, header in enumerate(header_row):
                if header:
                    header_lower = str(header).lower().replace(" ", "_").replace("(ru)", "ru")
                    # Handle variations
                    if header_lower in ["code"]:
                        header_map["code"] = col_idx
                    elif header_lower in ["name"]:
                        header_map["name"] = col_idx
                    elif header_lower in ["name_ru", "name_ru"]:
                        header_map["name_ru"] = col_idx
                    elif header_lower in ["description"]:
                        header_map["description"] = col_idx
                    elif header_lower in ["unit"]:
                        header_map["unit"] = col_idx
                    elif header_lower in ["min_value", "min"]:
                        header_map["min_value"] = col_idx
                    elif header_lower in ["max_value", "max"]:
                        header_map["max_value"] = col_idx
                    elif header_lower in ["active"]:
                        header_map["active"] = col_idx
                    elif header_lower in ["category_code", "category"]:
                        header_map["category_code"] = col_idx

            # Parse data rows
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or all(cell is None for cell in row):
                    continue

                metric_data: dict[str, Any] = {}

                for field, col_idx in header_map.items():
                    if col_idx < len(row):
                        value = row[col_idx]
                        if value is not None:
                            if field == "active":
                                metric_data[field] = self._parse_bool(value)
                            elif field in ["min_value", "max_value"]:
                                metric_data[field] = value
                            else:
                                metric_data[field] = str(value).strip() if value else None

                if metric_data.get("code"):
                    metrics.append(metric_data)

            wb.close()

        except Exception as e:
            errors.append(ImportError(row=0, error=f"Failed to parse Excel file: {str(e)}"))

        return metrics, errors

    def _parse_json(self, content: bytes) -> tuple[list[dict[str, Any]], list[ImportError]]:
        """
        Parse JSON file content.

        Args:
            content: JSON file bytes

        Returns:
            Tuple of (parsed metrics list, parsing errors)
        """
        metrics: list[dict[str, Any]] = []
        errors: list[ImportError] = []

        try:
            data = json.loads(content.decode("utf-8"))

            # Support both {"metrics": [...]} and plain [...]
            if isinstance(data, list):
                raw_metrics = data
            elif isinstance(data, dict) and "metrics" in data:
                raw_metrics = data["metrics"]
            else:
                errors.append(ImportError(row=0, error="Invalid JSON structure: expected array or {'metrics': [...]}"))
                return metrics, errors

            for idx, item in enumerate(raw_metrics, start=2):
                if not isinstance(item, dict):
                    errors.append(ImportError(row=idx, error=f"Invalid item at index {idx}: expected object"))
                    continue

                metric_data: dict[str, Any] = {}

                for field in ["code", "name", "name_ru", "description", "unit", "category_code"]:
                    if field in item and item[field] is not None:
                        metric_data[field] = str(item[field]).strip()

                for field in ["min_value", "max_value"]:
                    if field in item and item[field] is not None:
                        metric_data[field] = item[field]

                if "active" in item:
                    metric_data["active"] = self._parse_bool(item["active"])

                if metric_data.get("code"):
                    metrics.append(metric_data)

        except json.JSONDecodeError as e:
            errors.append(ImportError(row=0, error=f"Invalid JSON: {str(e)}"))
        except Exception as e:
            errors.append(ImportError(row=0, error=f"Failed to parse JSON file: {str(e)}"))

        return metrics, errors

    def _detect_changes(
        self,
        existing,
        new_data: dict[str, Any],
        existing_category_code: str | None = None,
    ) -> dict[str, str] | None:
        """
        Detect changes between existing metric and new data.

        Args:
            existing: Existing MetricDef object
            new_data: New metric data from import
            existing_category_code: Current category code of the metric (optional)

        Returns:
            Dictionary of changes or None if no changes
        """
        changes: dict[str, str] = {}

        field_mappings = [
            ("name", "name"),
            ("name_ru", "name_ru"),
            ("description", "description"),
            ("unit", "unit"),
        ]

        for attr, key in field_mappings:
            old_val = getattr(existing, attr, None)
            new_val = new_data.get(key)

            if old_val != new_val and new_val is not None:
                old_display = str(old_val) if old_val else "(empty)"
                new_display = str(new_val) if new_val else "(empty)"
                changes[key] = f"{old_display} -> {new_display}"

        # Handle category_code changes
        new_category_code = new_data.get("category_code")
        if new_category_code is not None:
            new_category_code = new_category_code.strip() if new_category_code else None
        # Detect change only when new_category_code is explicitly provided in import data
        if "category_code" in new_data:
            if existing_category_code != new_category_code:
                old_display = existing_category_code if existing_category_code else "(none)"
                new_display = new_category_code if new_category_code else "(none)"
                changes["category_code"] = f"{old_display} -> {new_display}"

        # Handle numeric fields
        for field in ["min_value", "max_value"]:
            old_val = getattr(existing, field, None)
            new_val = new_data.get(field)

            if new_val is not None:
                new_decimal = self._to_decimal(new_val)
                if old_val != new_decimal:
                    old_display = str(old_val) if old_val is not None else "(empty)"
                    new_display = str(new_decimal) if new_decimal is not None else "(empty)"
                    changes[field] = f"{old_display} -> {new_display}"

        # Handle active field
        if "active" in new_data:
            new_active = new_data["active"]
            if existing.active != new_active:
                changes["active"] = f"{existing.active} -> {new_active}"

        return changes if changes else None

    @staticmethod
    def _parse_bool(value: Any) -> bool:
        """Parse boolean value from various formats."""
        if isinstance(value, bool):
            return value
        if isinstance(value, str):
            return value.lower() in ("true", "1", "yes", "y", "on")
        if isinstance(value, (int, float)):
            return bool(value)
        return True

    @staticmethod
    def _to_decimal(value: Any) -> Decimal | None:
        """Convert value to Decimal or None."""
        if value is None:
            return None
        if isinstance(value, Decimal):
            return value
        try:
            return Decimal(str(value))
        except (ValueError, TypeError):
            return None

    async def _resolve_category_id(
        self,
        category_code: str | None,
        category_cache: dict[str, Any],
    ) -> Any:
        """
        Resolve category_code to category_id, creating category if needed.

        Args:
            category_code: Category code from import data (can be None)
            category_cache: Cache dict {code: category} to avoid duplicate lookups/creates

        Returns:
            UUID of the category or None if category_code is None/empty
        """
        if not category_code or not category_code.strip():
            return None

        category_code = category_code.strip()

        # Check cache first
        if category_code in category_cache:
            return category_cache[category_code].id

        # Try to find existing category
        category = await self.category_repo.get_by_code(category_code)

        if not category:
            # Auto-create category with code as name
            try:
                category = await self.category_repo.create(
                    code=category_code,
                    name=category_code,  # Use code as name for auto-created categories
                    description="Auto-created during metric import",
                    sort_order=0,
                )
            except Exception:
                # Category might have been created by another concurrent request
                category = await self.category_repo.get_by_code(category_code)
                if not category:
                    return None

        # Cache and return
        category_cache[category_code] = category
        return category.id


def get_metric_import_export_service(db: AsyncSession) -> MetricImportExportService:
    """Factory function for MetricImportExportService."""
    return MetricImportExportService(db)
